"""
handlers_admin.py — Админ-панель через inline-кнопки.
Без сеансов. Каждый вопрос запускается отдельно.
"""

import os
import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from config import ADMIN_ID, DATA_DIR
import database as db

logger = logging.getLogger(__name__)

admin_input = {}


def is_admin(uid):
    return uid == ADMIN_ID


# ═══ ГЛАВНОЕ МЕНЮ ═══

async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        await update.message.reply_text("⛔ Нет прав.")
        return
    await show_main_menu(update.effective_chat.id, context)


async def show_main_menu(chat_id, context, msg_id=None):
    members = db.get_members()
    meetings = db.get_meetings()
    active_q = sum(1 for mt in meetings for q in mt["questions"] if q["status"] == "voting")

    text = (
        "🏛 *Панель секретаря*\n\n"
        f"👥 Участников: {len(members)}\n"
        f"📋 Заседаний: {len(meetings)}\n"
        f"🗳 Активных голосований: {active_q}"
    )
    kb = [
        [InlineKeyboardButton("👥 Участники", callback_data="m:members"),
         InlineKeyboardButton("📋 Заседания", callback_data="m:meetings")],
        [InlineKeyboardButton("📎 Загрузить xlsx", callback_data="m:upload"),
         InlineKeyboardButton("📥 Шаблон xlsx", callback_data="m:sample")],
    ]
    if msg_id:
        await context.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
            text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")
    else:
        await context.bot.send_message(chat_id=chat_id,
            text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


# ═══ УЧАСТНИКИ ═══

async def show_members(chat_id, ctx, msg_id):
    members = db.get_members()
    if not members:
        text = "👥 *Участники*\n\nПусто. Загрузите xlsx или добавьте вручную."
    else:
        lines = ["👥 *Участники*\n"]
        for i, m in enumerate(members, 1):
            s = "✅" if m["telegram_id"] else "⏳"
            lines.append(f"{i}. {s} {m['name']} `({m['pin']})`")
        lines.append(f"\n✅ в боте · ⏳ ожидание")
        text = "\n".join(lines)

    kb = [[InlineKeyboardButton("➕ Добавить", callback_data="mem:add")]]
    row = []
    for m in members:
        row.append(InlineKeyboardButton(f"❌ {m['name'][:12]}", callback_data=f"mem:del:{m['pin']}"))
        if len(row) == 2:
            kb.append(row)
            row = []
    if row:
        kb.append(row)
    kb.append([InlineKeyboardButton("◀️ Назад", callback_data="m:main")])

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


# ═══ ЗАСЕДАНИЯ ═══

async def show_meetings(chat_id, ctx, msg_id):
    meetings = db.get_meetings()
    if not meetings:
        text = "📋 *Заседания*\n\nНет заседаний."
    else:
        lines = ["📋 *Заседания*\n"]
        for mt in meetings:
            label = db.type_label(mt["type"])
            active = sum(1 for q in mt["questions"] if q["status"] == "voting")
            closed = sum(1 for q in mt["questions"] if q["status"] == "closed")
            icon = "🟢" if active > 0 else "📝"
            lines.append(f"{icon} {label} №{mt['protocol_number']} — {mt['date']}")
            lines.append(f"   Вопросов: {len(mt['questions'])} (завершено: {closed})")
        text = "\n".join(lines)

    kb = [
        [InlineKeyboardButton("📗 Илмий семинар", callback_data="mtg:new:seminar"),
         InlineKeyboardButton("📘 Илмий кенгаш", callback_data="mtg:new:council")],
    ]
    for mt in meetings:
        label = db.type_label(mt["type"])
        kb.append([InlineKeyboardButton(
            f"{'📗' if mt['type']=='seminar' else '📘'} №{mt['protocol_number']} ({mt['date']})",
            callback_data=f"mtg:open:{mt['id']}"
        )])
    kb.append([InlineKeyboardButton("◀️ Назад", callback_data="m:main")])

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


# ═══ ЗАСЕДАНИЕ — ДЕТАЛИ ═══

async def show_meeting(chat_id, ctx, msg_id, mid):
    mt = db.get_meeting(mid)
    if not mt:
        return

    label = db.type_label(mt["type"])
    members = db.get_members()
    active_q = sum(1 for q in mt["questions"] if q["status"] == "voting")

    text = (
        f"{'📗' if mt['type']=='seminar' else '📘'} *{label}*\n"
        f"📅 {mt['date']} · Протокол №{mt['protocol_number']}\n"
        f"👥 Присутствующих: {len(mt['attendees'])} из {len(members)}\n"
        f"❓ Вопросов: {len(mt['questions'])}"
    )
    if active_q:
        text += f"\n🟢 Активных: {active_q}"

    kb = [
        [InlineKeyboardButton(f"👥 Присутствующие ({len(mt['attendees'])})", callback_data=f"att:{mid}:0")],
        [InlineKeyboardButton("❓ Вопросы", callback_data=f"qst:{mid}"),
         InlineKeyboardButton("➕ Вопрос", callback_data=f"q:add:{mid}")],
        [InlineKeyboardButton("📄 Отчёт DOCX", callback_data=f"mtg:rep:{mid}")],
        [InlineKeyboardButton("🗑 Удалить", callback_data=f"mtg:del:{mid}"),
         InlineKeyboardButton("◀️ Назад", callback_data="m:meetings")],
    ]

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


# ═══ ПРИСУТСТВУЮЩИЕ ═══

async def show_attendees(chat_id, ctx, msg_id, mid, page=0):
    mt = db.get_meeting(mid)
    if not mt:
        return
    members = db.get_members()
    per = 8
    pages = max(1, (len(members) + per - 1) // per)
    page = min(page, pages - 1)
    chunk = members[page * per:(page + 1) * per]

    text = (
        f"👥 *Присутствующие* ({len(mt['attendees'])}/{len(members)})\n"
        f"Стр. {page+1}/{pages} · Нажмите для ✅/⬜"
    )

    kb = []
    for m in chunk:
        ch = "✅" if m["pin"] in mt["attendees"] else "⬜"
        kb.append([InlineKeyboardButton(f"{ch} {m['name']}", callback_data=f"att:t:{mid}:{m['pin']}:{page}")])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️", callback_data=f"att:{mid}:{page-1}"))
    nav.append(InlineKeyboardButton("Все ✅", callback_data=f"att:all:{mid}:{page}"))
    if page < pages - 1:
        nav.append(InlineKeyboardButton("➡️", callback_data=f"att:{mid}:{page+1}"))
    kb.append(nav)
    kb.append([InlineKeyboardButton("◀️ К заседанию", callback_data=f"mtg:open:{mid}")])

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


# ═══ ВОПРОСЫ ═══

async def show_questions(chat_id, ctx, msg_id, mid):
    mt = db.get_meeting(mid)
    if not mt:
        return

    label = db.type_label(mt["type"])

    if not mt["questions"]:
        text = f"❓ *Вопросы — {label} №{mt['protocol_number']}*\n\nНет вопросов."
    else:
        lines = [f"❓ *Вопросы — {label} №{mt['protocol_number']}*\n"]
        for i, q in enumerate(mt["questions"], 1):
            v = q["votes"]
            total = v["for"] + v["against"] + v["abstain"]
            icon = {"draft": "📝", "voting": "🟢", "closed": "🔴"}[q["status"]]

            lines.append(f"{icon} *{i}. {q['text']}*")
            if q["status"] == "closed":
                bar = _result_bar(v)
                verdict = "✔️ Принято" if v["for"] > v["against"] else "✖️ Не принято" if v["for"] < v["against"] else "➖ Равны"
                lines.append(f"   {bar}")
                lines.append(f"   ✅{v['for']} ❌{v['against']} ⬜{v['abstain']} → {verdict}")
            elif q["status"] == "voting":
                lines.append(f"   🗳 Голосуют... ({len(q['voted_pins'])}/{len(mt['attendees'])})")
            else:
                lines.append(f"   ⏳ Ожидает запуска")
            lines.append("")
        text = "\n".join(lines)

    kb = [[InlineKeyboardButton("➕ Добавить вопрос", callback_data=f"q:add:{mid}")]]
    for q in mt["questions"]:
        icon = {"draft": "📝", "voting": "🟢", "closed": "🔴"}[q["status"]]
        kb.append([InlineKeyboardButton(
            f"{icon} {q['text'][:30]}...",
            callback_data=f"q:open:{mid}:{q['id']}"
        )])
    kb.append([InlineKeyboardButton("◀️ К заседанию", callback_data=f"mtg:open:{mid}")])

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


def _result_bar(v):
    """Визуальная полоска результатов."""
    total = v["for"] + v["against"] + v["abstain"]
    if total == 0:
        return "░░░░░░░░░░ нет голосов"
    width = 10
    f = round(v["for"] / total * width)
    a = round(v["against"] / total * width)
    ab = width - f - a
    return "🟩" * f + "🟥" * a + "⬜" * ab


# ═══ ВОПРОС — ДЕТАЛИ ═══

async def show_question(chat_id, ctx, msg_id, mid, qid):
    mt = db.get_meeting(mid)
    q = db.get_question(mid, qid)
    if not mt or not q:
        return

    v = q["votes"]
    total = v["for"] + v["against"] + v["abstain"]
    status_text = {"draft": "📝 Ожидает", "voting": "🟢 Идёт голосование", "closed": "🔴 Завершён"}[q["status"]]

    lines = [
        f"❓ *{q['text']}*\n",
        f"Статус: {status_text}",
    ]

    if q["status"] in ("voting", "closed"):
        lines.append(f"\n📊 *Результаты:*")
        lines.append(f"   {_result_bar(v)}")
        lines.append(f"   ✅ За: *{v['for']}*")
        lines.append(f"   ❌ Против: *{v['against']}*")
        lines.append(f"   ⬜ Воздержался: *{v['abstain']}*")
        lines.append(f"   📊 Всего: {total} из {len(mt['attendees'])}")
        if q["status"] == "closed":
            if v["for"] > v["against"]:
                lines.append(f"\n   ✔️ *РЕШЕНИЕ ПРИНЯТО*")
            elif v["for"] < v["against"]:
                lines.append(f"\n   ✖️ *РЕШЕНИЕ НЕ ПРИНЯТО*")
            else:
                lines.append(f"\n   ➖ *ГОЛОСА РАВНЫ*")

    text = "\n".join(lines)

    kb = []
    if q["status"] == "draft":
        if mt["attendees"]:
            kb.append([InlineKeyboardButton("🗳 Запустить голосование", callback_data=f"q:start:{mid}:{qid}")])
        else:
            kb.append([InlineKeyboardButton("⚠️ Сначала отметьте присутствующих", callback_data=f"att:{mid}:0")])
        kb.append([InlineKeyboardButton("🗑 Удалить", callback_data=f"q:del:{mid}:{qid}")])

    elif q["status"] == "voting":
        not_voted = len(mt["attendees"]) - len(q["voted_pins"])
        kb.append([InlineKeyboardButton(f"⏹ Остановить (не голос.: {not_voted})", callback_data=f"q:stop:{mid}:{qid}")])

    elif q["status"] == "closed":
        kb.append([InlineKeyboardButton("📄 Отчёт", callback_data=f"mtg:rep:{mid}")])

    kb.append([InlineKeyboardButton("◀️ К вопросам", callback_data=f"qst:{mid}")])

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=text, reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown")


# ═══ CALLBACK ROUTER ═══

async def handle_admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if not is_admin(query.from_user.id):
        await query.answer("⛔", show_alert=True)
        return
    await query.answer()

    d = query.data
    cid = query.message.chat_id
    mid_msg = query.message.message_id

    # Меню
    if d == "m:main": await show_main_menu(cid, context, mid_msg)
    elif d == "m:members": await show_members(cid, context, mid_msg)
    elif d == "m:meetings": await show_meetings(cid, context, mid_msg)
    elif d == "m:upload":
        admin_input[ADMIN_ID] = {"action": "upload"}
        await context.bot.edit_message_text(chat_id=cid, message_id=mid_msg,
            text="📎 Отправьте .xlsx файл\n(A=ФИО, B=PIN)\n\n/cancel — отмена", parse_mode="Markdown")
    elif d == "m:sample":
        await send_sample(cid, context)

    # Участники
    elif d == "mem:add":
        admin_input[ADMIN_ID] = {"action": "add_member"}
        await context.bot.edit_message_text(chat_id=cid, message_id=mid_msg,
            text="➕ Отправьте:\n`ФИО, PIN`\n\nПример: `Иванов Иван, 1001`\n\n/cancel", parse_mode="Markdown")
    elif d.startswith("mem:del:"):
        pin = d.split(":")[2]
        db.remove_member(pin)
        await query.answer("Удалён", show_alert=True)
        await show_members(cid, context, mid_msg)

    # Заседания
    elif d.startswith("mtg:new:"):
        mtype = d.split(":")[2]
        admin_input[ADMIN_ID] = {"action": "create_meeting", "type": mtype}
        label = db.type_label(mtype)
        await context.bot.edit_message_text(chat_id=cid, message_id=mid_msg,
            text=f"➕ *{label}*\n\nОтправьте:\n`ДД.ММ.ГГГГ, номер`\n\nПример: `20.03.2026, 5`\n\n/cancel", parse_mode="Markdown")
    elif d.startswith("mtg:open:"):
        await show_meeting(cid, context, mid_msg, d.split(":")[2])
    elif d.startswith("mtg:del:"):
        db.delete_meeting(d.split(":")[2])
        await query.answer("Удалено", show_alert=True)
        await show_meetings(cid, context, mid_msg)
    elif d.startswith("mtg:rep:"):
        await send_report(cid, context, d.split(":")[2])

    # Присутствующие
    elif d.startswith("att:t:"):
        p = d.split(":")
        db.toggle_attendee(p[2], p[3])
        await show_attendees(cid, context, mid_msg, p[2], int(p[4]))
    elif d.startswith("att:all:"):
        p = d.split(":")
        db.set_all_attendees(p[2])
        await show_attendees(cid, context, mid_msg, p[2], int(p[3]))
    elif d.startswith("att:"):
        p = d.split(":")
        await show_attendees(cid, context, mid_msg, p[1], int(p[2]))

    # Вопросы
    elif d.startswith("qst:"):
        await show_questions(cid, context, mid_msg, d.split(":")[1])
    elif d.startswith("q:add:"):
        mid = d.split(":")[2]
        admin_input[ADMIN_ID] = {"action": "add_question", "mid": mid}
        await context.bot.edit_message_text(chat_id=cid, message_id=mid_msg,
            text="❓ Отправьте текст вопроса\n\nПример:\n`Утвердить оппонентов Иванова И.И.`\n\n/cancel", parse_mode="Markdown")
    elif d.startswith("q:open:"):
        p = d.split(":")
        await show_question(cid, context, mid_msg, p[2], p[3])
    elif d.startswith("q:del:"):
        p = d.split(":")
        db.delete_question(p[2], p[3])
        await query.answer("Удалён", show_alert=True)
        await show_questions(cid, context, mid_msg, p[2])
    elif d.startswith("q:start:"):
        p = d.split(":")
        await start_voting(cid, context, mid_msg, p[2], p[3])
    elif d.startswith("q:stop:"):
        p = d.split(":")
        await stop_voting(cid, context, mid_msg, p[2], p[3])


# ═══ ГОЛОСОВАНИЕ ═══

async def start_voting(chat_id, ctx, msg_id, mid, qid):
    mt = db.get_meeting(mid)
    q = db.get_question(mid, qid)
    if not mt or not q:
        return
    if not mt["attendees"]:
        await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id, text="❌ Отметьте присутствующих!")
        return

    db.start_question_voting(mid, qid)

    # Уведомить
    notified = 0
    for m in db.get_members():
        if m["pin"] in mt["attendees"] and m["telegram_id"]:
            try:
                kb = [[
                    InlineKeyboardButton("✅ За", callback_data=f"v:{mid}:{qid}:for"),
                    InlineKeyboardButton("❌ Против", callback_data=f"v:{mid}:{qid}:against"),
                    InlineKeyboardButton("⬜ Воздерж.", callback_data=f"v:{mid}:{qid}:abstain"),
                ]]
                label = db.type_label(mt["type"])
                await ctx.bot.send_message(
                    chat_id=m["telegram_id"],
                    text=f"🗳 *Голосование!*\n\n{label} №{mt['protocol_number']}\n📅 {mt['date']}\n\n❓ *{q['text']}*",
                    reply_markup=InlineKeyboardMarkup(kb), parse_mode="Markdown"
                )
                notified += 1
            except Exception as e:
                logger.warning(f"Не уведомлён {m['name']}: {e}")

    await ctx.bot.edit_message_text(chat_id=chat_id, message_id=msg_id,
        text=f"🗳 *Голосование запущено!*\n\n❓ {q['text']}\n📨 Уведомлено: {notified}/{len(mt['attendees'])}",
        parse_mode="Markdown")

    # Показать детали
    msg = await ctx.bot.send_message(chat_id=chat_id, text="⏳")
    await show_question(chat_id, ctx, msg.message_id, mid, qid)


async def stop_voting(chat_id, ctx, msg_id, mid, qid):
    db.stop_question_voting(mid, qid)
    mt = db.get_meeting(mid)
    q = db.get_question(mid, qid)

    for m in db.get_members():
        if m["pin"] in mt["attendees"] and m["telegram_id"]:
            try:
                await ctx.bot.send_message(chat_id=m["telegram_id"],
                    text=f"🏁 Голосование завершено: \"{q['text'][:40]}...\"")
            except Exception:
                pass

    await show_question(chat_id, ctx, msg_id, mid, qid)


# ═══ ТЕКСТОВЫЙ ВВОД ═══

async def handle_admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not is_admin(uid):
        return False
    state = admin_input.get(uid)
    if not state:
        return False

    text = update.message.text.strip()
    if text.startswith("/"):
        admin_input.pop(uid, None)
        return False

    action = state["action"]
    admin_input.pop(uid, None)

    if action == "add_member":
        parts = text.split(",", 1)
        if len(parts) != 2:
            await update.message.reply_text("❌ Формат: `ФИО, PIN`\n/admin", parse_mode="Markdown")
            return True
        name, pin = parts[0].strip(), parts[1].strip()
        if db.add_member(name, pin):
            await update.message.reply_text(f"✅ {name} (PIN: {pin})\n/admin")
        else:
            await update.message.reply_text(f"❌ PIN {pin} занят\n/admin")

    elif action == "create_meeting":
        parts = text.split(",", 1)
        if len(parts) != 2:
            await update.message.reply_text("❌ Формат: `ДД.ММ.ГГГГ, номер`\n/admin", parse_mode="Markdown")
            return True
        date_str, proto = parts[0].strip(), parts[1].strip()
        mtype = state.get("type", "seminar")
        db.create_meeting(mtype, date_str, proto)
        label = db.type_label(mtype)
        await update.message.reply_text(f"✅ {label}: {date_str}, №{proto}\n/admin")

    elif action == "add_question":
        mid = state["mid"]
        db.add_question(mid, text)
        await update.message.reply_text(f"✅ Вопрос добавлен\n/admin")

    elif action == "upload":
        await update.message.reply_text("📎 Отправьте *файл* .xlsx\n/admin", parse_mode="Markdown")

    return True


# ═══ ФАЙЛ ═══

async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return

    doc = update.message.document
    if not (doc.file_name or "").lower().endswith(".xlsx"):
        await update.message.reply_text("❌ Нужен .xlsx\n/sample")
        return

    admin_input.pop(ADMIN_ID, None)
    os.makedirs(DATA_DIR, exist_ok=True)
    fp = os.path.join(DATA_DIR, "upload.xlsx")
    f = await context.bot.get_file(doc.file_id)
    await f.download_to_drive(fp)

    try:
        from openpyxl import load_workbook
        wb = load_workbook(fp)
        ws = wb.active
        added, skipped = 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            name = str(row[0]).strip() if row[0] else ""
            pin = str(row[1]).strip() if row[1] else ""
            if pin.endswith(".0"):
                pin = pin[:-2]
            if not name or not pin:
                continue
            if db.add_member(name, pin):
                added += 1
            else:
                skipped += 1
        await update.message.reply_text(
            f"📊 ✅ Добавлено: {added} · ⏭ Пропущено: {skipped}\n👥 Всего: {len(db.get_members())}\n/admin",
            parse_mode="Markdown")
    except Exception as e:
        await update.message.reply_text(f"❌ {e}\n/sample")
    finally:
        if os.path.exists(fp):
            os.remove(fp)


async def cmd_sample(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    await send_sample(update.effective_chat.id, context)


async def send_sample(chat_id, ctx):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Участники"
    hf = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    hfill = PatternFill(start_color="0D6E6E", end_color="0D6E6E", fill_type="solid")
    b = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
    for col, h in enumerate(["ФИО", "PIN"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = hf, hfill, b
    for r, (n, p) in enumerate([("Иванов Иван Иванович", "1001"), ("Петрова Мария Сергеевна", "1002")], 2):
        ws.cell(row=r, column=1, value=n).border = b
        ws.cell(row=r, column=2, value=p).border = b
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    os.makedirs(DATA_DIR, exist_ok=True)
    fp = os.path.join(DATA_DIR, "shablon.xlsx")
    wb.save(fp)
    with open(fp, "rb") as f:
        await ctx.bot.send_document(chat_id=chat_id, document=f, filename="shablon.xlsx",
            caption="A=ФИО, B=PIN. Строку 1 не удалять.")
    os.remove(fp)


# ═══ ОТЧЁТ ═══

async def send_report(chat_id, ctx, mid):
    from report import generate_report
    fp = generate_report(mid)
    if fp:
        with open(fp, "rb") as f:
            await ctx.bot.send_document(chat_id=chat_id, document=f,
                filename=fp.split("/")[-1], caption="📄 Отчёт для протокола")
    else:
        await ctx.bot.send_message(chat_id=chat_id, text="❌ Ошибка отчёта")
